from openpyxl import load_workbook
from . import db
from .models import CNE, comuna

NO_COMUNAS = 0
NUMERO_FORMULARIOS = 2

def retrieve_comunas_list():
    comunas_list = []

    # Load workbook
    wb = load_workbook('regionesComunas.xlsx')
    
    # Get sheets
    comunas_sheet = wb['comunas']
    regiones_sheet = wb['regiones']

    # Dictionary to store region names based on id_region
    region_id_name_mapping = {region_row[0]: region_row[1] for region_row in regiones_sheet.iter_rows(values_only=True)}

    # Iterate through "comunas" sheet
    for id_comuna, descripcion, id_region in comunas_sheet.iter_rows(min_row=2, values_only=True):
        nombre_region = region_id_name_mapping.get(id_region, '')
        comunas_list.append(comuna(codigo_comuna=id_comuna, nombre_comuna=descripcion, codigo_region=id_region, nombre_region=nombre_region))
    
    return comunas_list

def retrieve_formularios_list():
    formularios = [
        CNE(codigo_cne='8', nombre_cne='Compraventa'),
        CNE(codigo_cne='99', nombre_cne='Regularización del Patrimonio'),
        # Add formularios as needed
    ]
    return formularios

def seed_database():
    db_seeded = False
    print("Running seeder...")

    global NO_COMUNAS
    global NUMERO_FORMULARIOS
    
    if db.session.query(comuna).count() == NO_COMUNAS:
        print("Seeding database with comunas...")

        comunas_list = retrieve_comunas_list()
        for comuna_obj in comunas_list:
            db.session.add(comuna_obj)
        db_seeded = True
    
    if db.session.query(CNE).count() < NUMERO_FORMULARIOS:
        print("Seeding database with formularios...")
        formularios_list = retrieve_formularios_list()
        for formulario_obj in formularios_list:
            db.session.add(formulario_obj)
        db_seeded = True

    if db_seeded:
        db.session.commit()
        print("Database seeded successfully")
    else:
        print("Database is already seeded")

